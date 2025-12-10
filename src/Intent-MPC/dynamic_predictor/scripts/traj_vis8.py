#!/usr/bin/env python3
import rospy
import math
import time
import os
import numpy as np
from datetime import datetime
from visualization_msgs.msg import MarkerArray, Marker
from std_msgs.msg import Float64MultiArray
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font

class TrajectoryEvaluator:
    def __init__(self):
        self.pred_trajectories = defaultdict(dict)  # {obstacle_id: {intent_type: [points]}}
        self.history_trajectories = defaultdict(list)  # {obstacle_id: [points]}
        self.intent_probabilities = defaultdict(lambda: [0.25, 0.25, 0.25, 0.25])  # {obstacle_id: [P_forward, P_left, P_right, P_stop]}
        self.adaptive_metrics = {}  # {obstacle_id: {"D_t": value, "s_adaptive": value}}
        self.last_pred_time = 0
        self.last_history_time = 0
        self.all_evaluation_data = []
        
        # 预测参数（与 C++ 保持一致）
        self.dt = 0.1  # 预测时间步长
        self.total_pred_steps = int(3.0 / self.dt)  # 3秒预测，共30步
        
        # 自适应权重参数（从 ROS 参数服务器读取，与 C++ 保持一致）
        self.lambda1 = rospy.get_param('/dynamic_predictor/lambda1', 1.0)
        self.lambda2 = rospy.get_param('/dynamic_predictor/lambda2', 2.0)
        self.M_thresh = rospy.get_param('/dynamic_predictor/M_thresh', 1.0)
        self.s_max = rospy.get_param('/dynamic_predictor/s_max', 2.0)
        self.gamma1 = rospy.get_param('/dynamic_predictor/gamma1', 0.5)  # 默认值
        self.gamma2 = rospy.get_param('/dynamic_predictor/gamma2', 0.5)  # 默认值
        
        # 获取桌面路径
        self.desktop_path = self.get_desktop_path()
        rospy.loginfo(f"检测到桌面路径: {self.desktop_path}")
        
        # 订阅话题
        self.pred_sub = rospy.Subscriber('/dynamic_predictor/predict_trajectories',
                                        MarkerArray, self.pred_callback)
        self.history_sub = rospy.Subscriber('/dynamic_predictor/history_trajectories',
                                           MarkerArray, self.history_callback)
        self.intent_sub = rospy.Subscriber('/dynamic_predictor/intent_probability',
                                           MarkerArray, self.intent_callback)
        self.adaptive_sub = rospy.Subscriber('/dynamic_predictor/adaptive_metrics',
                                           Float64MultiArray, self.adaptive_callback)
        
        # 定时评估
        self.eval_timer = rospy.Timer(rospy.Duration(0.1), self.evaluate)
        
        # 确保程序退出时保存文件
        rospy.on_shutdown(self.save_final_excel)
        print("轨迹评估器已启动，按 Ctrl+C 时自动保存文件到桌面")

    def get_desktop_path(self):
        """自动检测桌面路径（兼容中英文系统）"""
        home = os.path.expanduser("~")
        chinese_desktop = os.path.join(home, "桌面")
        english_desktop = os.path.join(home, "Desktop")
        if os.path.exists(chinese_desktop) and os.path.isdir(chinese_desktop):
            return chinese_desktop
        if os.path.exists(english_desktop) and os.path.isdir(english_desktop):
            return english_desktop
        rospy.logwarn("未找到桌面目录，使用家目录")
        return home

    def pred_callback(self, msg):
        """接收预测轨迹"""
        self.last_pred_time = time.time()
        # marker.id 的编码：对于障碍物 i，意图 j，marker.id = i * numIntent_ + j
        # numIntent_ = 4 (forward, left, right, stop)
        num_intent = 4
        for marker in msg.markers:
            obstacle_id = marker.id // num_intent
            intent_type = marker.id % num_intent
            # 确保 intent_type 在有效范围内 (0-3)
            if intent_type >= 0 and intent_type < num_intent:
                if marker.type == Marker.LINE_STRIP:
                    points = [[p.x, p.y, p.z] for p in marker.points]
                    self.pred_trajectories[obstacle_id][intent_type] = points

    def history_callback(self, msg):
        """接收历史轨迹"""
        self.last_history_time = time.time()
        for marker in msg.markers:
            obstacle_id = marker.id
            if marker.type == Marker.LINE_STRIP:
                points = [[p.x, p.y, p.z] for p in marker.points]
                self.history_trajectories[obstacle_id] = points

    def intent_callback(self, msg):
        """接收意图概率（注意：C++ 只发布最大概率意图的文本，无法获取完整分布）"""
        # 由于 C++ 的 publishIntentVis 只发布最大概率意图的文本，
        # 我们无法直接获取完整的意图概率分布
        # 这里保留回调函数以备将来扩展
        pass

    def adaptive_callback(self, msg):
        """接收自适应指标 D_t 和 s_adaptive（从 C++ 直接获取，更准确）"""
        # 数据格式：对于 N 个障碍物，数组长度为 2*N
        # 每个障碍物有两个值：[D_t, s_adaptive]
        data = msg.data
        num_obstacles = len(data) // 2
        
        for i in range(num_obstacles):
            obstacle_id = i
            idx = i * 2
            if idx + 1 < len(data):
                Dt = data[idx]
                s_adaptive = data[idx + 1]
                self.adaptive_metrics[obstacle_id] = {
                    "D_t": Dt,
                    "s_adaptive": s_adaptive
                }

    def calculate_2d_distance(self, p1, p2):
        """计算2D距离（忽略z分量，与C++保持一致）"""
        return math.sqrt((p1[0] - p2[0])**2 + (p1[1] - p2[1])**2)

    def compute_intent_entropy(self, intent_prob):
        """计算意图熵 H_t（与 C++ 逻辑一致）"""
        if len(intent_prob) == 0:
            return 0.0
        entropy = 0.0
        for p in intent_prob:
            if p > 1e-10:  # 避免 log(0)
                entropy -= p * math.log(p)
        return entropy

    def compute_jerk_norm(self, history_traj):
        """从历史轨迹计算加加速度幅值 ||j_t||（简化版本）"""
        if len(history_traj) < 3:
            return 0.0
        
        # 计算速度（2D）
        velocities = []
        for i in range(len(history_traj) - 1):
            vx = (history_traj[i+1][0] - history_traj[i][0]) / self.dt
            vy = (history_traj[i+1][1] - history_traj[i][1]) / self.dt
            velocities.append([vx, vy])
        
        if len(velocities) < 2:
            return 0.0
        
        # 计算加速度（2D）
        accelerations = []
        for i in range(len(velocities) - 1):
            ax = (velocities[i+1][0] - velocities[i][0]) / self.dt
            ay = (velocities[i+1][1] - velocities[i][1]) / self.dt
            accelerations.append([ax, ay])
        
        if len(accelerations) < 2:
            return 0.0
        
        # 计算加加速度（jerk）
        curr_acc = np.array(accelerations[-1])
        prev_acc = np.array(accelerations[-2])
        jerk = (curr_acc - prev_acc) / self.dt
        jerk_norm = np.linalg.norm(jerk)
        
        return jerk_norm

    def compute_nis_simplified(self, history_traj):
        """计算简化的 NIS（归一化残差平方）"""
        if len(history_traj) < 3:
            return 0.0
        
        # 简化的 CA 模型预测：假设恒定加速度
        # 使用最近两个点预测下一个点
        if len(history_traj) >= 2:
            prev_pos = np.array(history_traj[-2][:2])  # 2D
            curr_pos = np.array(history_traj[-1][:2])  # 2D
            
            # 计算速度
            vel = (curr_pos - prev_pos) / self.dt
            
            # CA 模型预测：假设恒定加速度，用最近的速度预测
            if len(history_traj) >= 3:
                prev_prev_pos = np.array(history_traj[-3][:2])
                prev_vel = (prev_pos - prev_prev_pos) / self.dt
                acc = (vel - prev_vel) / self.dt
                ca_pred_pos = curr_pos + vel * self.dt + 0.5 * acc * self.dt**2
            else:
                ca_pred_pos = curr_pos + vel * self.dt
            
            # 实际位置（如果有下一个点，否则用当前位置）
            if len(history_traj) >= 2:
                actual_pos = curr_pos
            else:
                actual_pos = curr_pos
            
            # 残差
            residual = actual_pos - ca_pred_pos
            
            # 简化的协方差矩阵（单位矩阵 * 0.1）
            residual_cov = np.eye(2) * 0.1
            residual_cov_inv = np.linalg.inv(residual_cov)
            
            # NIS
            nis = residual.T @ residual_cov_inv @ residual
            return nis
        
        return 0.0

    def compute_adaptive_s(self, Ht, Mt):
        """计算自适应权重 s_adaptive 和衰减因子 D_t（与 C++ 逻辑一致）"""
        # 计算熵约束项：e^(-λ1*Ht)
        D_entropy = math.exp(-self.lambda1 * Ht)
        
        # 计算运动诊断约束项：1/(1 + e^(λ2*(Mt - M_thresh)))
        D_motion = 1.0 / (1.0 + math.exp(self.lambda2 * (Mt - self.M_thresh)))
        
        # 衰减因子 D_t
        Dt = D_entropy * D_motion
        
        # 自适应权重（限制在[1, s_max]）
        s_adaptive = 1.0 + (self.s_max - 1.0) * Dt
        s_adaptive = max(1.0, min(s_adaptive, self.s_max))  # 截断到有效范围
        
        return Dt, s_adaptive

    def calculate_ade_fde(self, pred_points, ref_points):
        """
        计算 ADE 和 FDE（与 C++ 逻辑一致）
        - 使用2D距离（忽略z分量）
        - 回测：用历史轨迹的最后 total_pred_steps 个点作为"未来"参考
        """
        if not pred_points or not ref_points:
            return float('inf'), float('inf')
        
        hist_size = len(ref_points)
        if hist_size < self.total_pred_steps + 1:
            # 历史轨迹不够长，无法进行回测
            return float('inf'), float('inf')
        
        # 回测：用历史轨迹的最后 total_pred_steps 个点作为"未来"参考
        # 参考 C++: refStartIdx = histSize - totalPredSteps
        ref_start_idx = hist_size - self.total_pred_steps
        valid_steps = min(self.total_pred_steps, len(pred_points))
        
        if valid_steps == 0:
            return float('inf'), float('inf')
        
        # 计算 ADE 和 FDE（使用2D距离）
        ade = 0.0
        fde = 0.0
        actual_steps = 0
        
        for t in range(valid_steps):
            ref_idx = ref_start_idx + t
            if ref_idx >= hist_size:
                break
            
            # 计算2D距离（忽略z分量）
            error_2d = self.calculate_2d_distance(pred_points[t], ref_points[ref_idx])
            ade += error_2d
            actual_steps += 1
            
            # FDE 是最后一步的误差
            if t == valid_steps - 1:
                fde = error_2d
        
        if actual_steps > 0:
            ade /= actual_steps
        else:
            return float('inf'), float('inf')
        
        return ade, fde

    def evaluate(self, event=None):
        """评估函数（与 C++ 的 calculateAndPrintErrors 逻辑一致）"""
        current_time = time.time()
        # 如果超过2秒没有收到数据，跳过
        if current_time - self.last_pred_time > 2.0 or current_time - self.last_history_time > 2.0:
            return
        
        # 遍历每个障碍物
        for obstacle_id in self.pred_trajectories:
            if obstacle_id not in self.history_trajectories:
                continue
            
            ref_traj = self.history_trajectories[obstacle_id]
            pred_trajs = self.pred_trajectories[obstacle_id]
            
            if not ref_traj or not pred_trajs:
                continue
            
            # 遍历所有意图的预测轨迹，找最小 ADE 和 FDE
            min_ade = float('inf')
            min_fde = float('inf')
            best_intent = -1
            
            for intent_type, pred_traj in pred_trajs.items():
                ade, fde = self.calculate_ade_fde(pred_traj, ref_traj)
                
                if ade < min_ade:
                    min_ade = ade
                    min_fde = fde
                    best_intent = intent_type
            
            # 从 C++ 直接获取 D_t 和 s_adaptive（更准确）
            if obstacle_id in self.adaptive_metrics:
                Dt = self.adaptive_metrics[obstacle_id]["D_t"]
                s_adaptive = self.adaptive_metrics[obstacle_id]["s_adaptive"]
            else:
                # 如果还没有收到 C++ 发布的值，尝试从历史轨迹计算（备用方案）
                try:
                    # 获取意图概率
                    intent_prob = self.intent_probabilities[obstacle_id]
                    
                    # 计算意图熵 H_t
                    Ht = self.compute_intent_entropy(intent_prob)
                    
                    # 计算运动诊断指标 M_t = gamma1 * NIS + gamma2 * ||j_t||
                    jerk_norm = self.compute_jerk_norm(ref_traj)
                    nis = self.compute_nis_simplified(ref_traj)
                    Mt = self.gamma1 * nis + self.gamma2 * jerk_norm
                    
                    # 计算 D_t 和 s_adaptive
                    Dt, s_adaptive = self.compute_adaptive_s(Ht, Mt)
                except Exception as e:
                    # 如果计算失败，使用默认值
                    rospy.logdebug(f"计算 D_t 和 s_adaptive 时出错 (障碍物 {obstacle_id}): {e}")
                    Dt = 0.0
                    s_adaptive = 1.0
            
            # 如果有有效结果，保存数据
            # 确保 best_intent 在有效范围内 (0-3)
            if best_intent != -1 and min_ade < float('inf') and best_intent < 4:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")
                self.all_evaluation_data.append({
                    "timestamp": timestamp,
                    "obstacle_id": obstacle_id,
                    "ADE": min_ade,
                    "FDE": min_fde,
                    "best_intent": best_intent,
                    "valid_steps": self.total_pred_steps,
                    "D_t": Dt,
                    "s_adaptive": s_adaptive
                })

    def save_final_excel(self):
        """保存结果到 Excel 文件（参考 traj_vis7.py）"""
        # 检查写入权限
        if not os.access(self.desktop_path, os.W_OK):
            rospy.logerr(f"目录不可写: {self.desktop_path}")
            self.desktop_path = "/tmp"
            rospy.logwarn(f"改为保存到: {self.desktop_path}")
        
        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"intent_eval_{timestamp}.xlsx"
        full_path = os.path.join(self.desktop_path, filename)
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "ADE_FDE评估"
        
        # 表头（与 C++ CSV 格式对应，添加 D_t 和 s_adaptive）
        headers = [
            "时间戳",
            "障碍物ID",
            "ADE(米)",
            "FDE(米)",
            "D_t",
            "s_adaptive",
            "最佳意图",
            "有效步数"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
        
        # 写入数据
        if self.all_evaluation_data:
            for row, data in enumerate(self.all_evaluation_data, 2):
                ws.cell(row=row, column=1, value=data["timestamp"])
                ws.cell(row=row, column=2, value=data["obstacle_id"])
                ws.cell(row=row, column=3, value=round(data["ADE"], 6))
                ws.cell(row=row, column=4, value=round(data["FDE"], 6))
                ws.cell(row=row, column=5, value=round(data.get("D_t", 0.0), 6))
                ws.cell(row=row, column=6, value=round(data.get("s_adaptive", 1.0), 6))
                ws.cell(row=row, column=7, value=data["best_intent"])
                ws.cell(row=row, column=8, value=data["valid_steps"])
        else:
            # 即使没有数据也创建文件
            ws.cell(row=2, column=1, value="无评估数据")
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
        
        # 保存文件
        try:
            wb.save(full_path)
            print(f"\n文件已保存到: {full_path}")
            print(f"你可以运行: xdg-open {full_path}")
        except Exception as e:
            print(f"保存失败: {e}")
            fallback = os.path.abspath(f"intent_eval_backup_{timestamp}.xlsx")
            try:
                wb.save(fallback)
                print(f"已保存到当前目录备用文件: {fallback}")
            except Exception as e2:
                print(f"备用保存也失败: {e2}")


if __name__ == '__main__':
    rospy.init_node('trajectory_evaluator_py')
    evaluator = TrajectoryEvaluator()
    try:
        rospy.spin()
    except KeyboardInterrupt:
        # 双重保障：即使 rospy.on_shutdown 失效，也手动调用保存
        evaluator.save_final_excel()

