#!/usr/bin/env python3
import rospy
import math
import time
from datetime import datetime
from visualization_msgs.msg import MarkerArray, Marker
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
# 想要输出excel但失败了

class SimpleEvaluator:
    def __init__(self):
        # 存储数据 {障碍物ID: {意图类型: 轨迹点}}
        self.pred_trajectories = defaultdict(dict)
        self.history_trajectories = defaultdict(list)
        
        # 存储时间戳用于对齐
        self.last_pred_time = 0
        self.last_history_time = 0
        
        # 新增：累积所有评估数据
        self.all_evaluation_data = []
        
        # 订阅话题
        self.pred_sub = rospy.Subscriber('/dynamic_predictor/predict_trajectories', 
                                        MarkerArray, self.pred_callback)
        self.history_sub = rospy.Subscriber('/dynamic_predictor/history_trajectories', 
                                           MarkerArray, self.history_callback)
        
        # 定时评估
        self.eval_timer = rospy.Timer(rospy.Duration(1.0), self.evaluate)
        
        print("简单轨迹评估器已启动，等待数据...")

    def pred_callback(self, msg):
        """预测轨迹回调"""
        self.last_pred_time = time.time()
        
        for marker in msg.markers:
            # 提取障碍物ID和意图类型
            obstacle_id = marker.id // 10  # 假设每10个ID是一个障碍物的不同意图
            intent_type = marker.id % 10
            
            if marker.type == Marker.LINE_STRIP:
                points = []
                for p in marker.points:
                    points.append([p.x, p.y, p.z])
                self.pred_trajectories[obstacle_id][intent_type] = points

    def history_callback(self, msg):
        """历史轨迹回调"""
        self.last_history_time = time.time()
        
        for marker in msg.markers:
            obstacle_id = marker.id
            
            if marker.type == Marker.LINE_STRIP:
                points = []
                for p in marker.points:
                    points.append([p.x, p.y, p.z])
                self.history_trajectories[obstacle_id] = points

    def calculate_distance(self, point1, point2):
        """计算两点间距离"""
        dx = point1[0] - point2[0]
        dy = point1[1] - point2[1]
        dz = point1[2] - point2[2]
        return math.sqrt(dx*dx + dy*dy + dz*dz)

    def calculate_ade_fde(self, pred_points, actual_points):
        """计算ADE和FDE"""
        if len(pred_points) == 0 or len(actual_points) == 0:
            return float('inf'), float('inf')
        
        # 取相同长度的轨迹
        min_length = min(len(pred_points), len(actual_points))
        
        # 计算ADE
        total_error = 0
        for i in range(min_length):
            distance = self.calculate_distance(pred_points[i], actual_points[i])
            total_error += distance
        ade = total_error / min_length
        
        # 计算FDE
        fde = self.calculate_distance(pred_points[min_length-1], actual_points[min_length-1])
        
        return ade, fde

    def evaluate(self, event=None):
        """评估函数（仅打印，不保存Excel，累积数据）"""
        # 检查数据是否新鲜（2秒内的数据）
        current_time = time.time()
        if current_time - self.last_pred_time > 2.0 or current_time - self.last_history_time > 2.0:
            return
        
        print("\n" + "="*60)
        print(f"轨迹预测评估结果 (时间: {datetime.now().strftime('%H:%M:%S')})")
        print("="*60)
        
        current_eval_data = []  # 当前评估周期的数据
        total_min_ade = 0
        total_min_fde = 0
        count = 0
        
        for obstacle_id in self.pred_trajectories:
            if obstacle_id in self.history_trajectories:
                actual_points = self.history_trajectories[obstacle_id]
                
                # 对每个意图计算ADE/FDE，取最好的
                min_ade = float('inf')
                min_fde = float('inf')
                best_intent = -1
                
                for intent_type, pred_points in self.pred_trajectories[obstacle_id].items():
                    ade, fde = self.calculate_ade_fde(pred_points, actual_points)
                    
                    if ade < min_ade:
                        min_ade = ade
                        min_fde = fde
                        best_intent = intent_type
                
                if min_ade < float('inf'):
                    # 实时打印
                    print(f"障碍物 {obstacle_id} (最佳意图:{best_intent}):")
                    print(f"  minADE: {min_ade:.3f}米")
                    print(f"  minFDE: {min_fde:.3f}米")
                    print(f"  轨迹长度: {len(actual_points)}个点")
                    print("-" * 40)
                    
                    # 累积当前周期数据
                    current_eval_data.append({
                        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "obstacle_id": obstacle_id,
                        "best_intent": best_intent,
                        "min_ade": min_ade,
                        "min_fde": min_fde,
                        "traj_length": len(actual_points)
                    })
                    
                    total_min_ade += min_ade
                    total_min_fde += min_fde
                    count += 1
        
        # 打印当前周期的平均值
        if count > 0:
            avg_ade = total_min_ade / count
            avg_fde = total_min_fde / count
            print(f"\n当前周期平均 (基于{count}个障碍物):")
            print(f"  平均minADE: {avg_ade:.3f}米")
            print(f"  平均minFDE: {avg_fde:.3f}米")
        else:
            print("没有找到匹配的轨迹数据进行评估")
        
        print("="*60)
        
        # 将当前周期数据添加到总列表
        self.all_evaluation_data.extend(current_eval_data)

    def save_final_excel(self):
        """程序结束时保存汇总Excel"""
        if not self.all_evaluation_data:
            print("\n没有评估数据可保存")
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "轨迹评估汇总"
        
        # 设置表头
        headers = ["时间戳", "障碍物ID", "最佳意图类型", "minADE(米)", "minFDE(米)", "轨迹长度(点)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # 写入所有累积数据
        for row, data in enumerate(self.all_evaluation_data, 2):
            ws.cell(row=row, column=1, value=data["timestamp"])
            ws.cell(row=row, column=2, value=data["obstacle_id"])
            ws.cell(row=row, column=3, value=data["best_intent"])
            ws.cell(row=row, column=4, value=round(data["min_ade"], 3))
            ws.cell(row=row, column=5, value=round(data["min_fde"], 3))
            ws.cell(row=row, column=6, value=data["traj_length"])
        
        # 计算并写入总体平均值
        total_ade = sum(data["min_ade"] for data in self.all_evaluation_data)
        total_fde = sum(data["min_fde"] for data in self.all_evaluation_data)
        total_count = len(self.all_evaluation_data)
        
        summary_row = total_count + 3
        cell = ws.cell(row=summary_row, column=1)
        cell.value = "总体平均值"
        cell.font = Font(bold=True)
        
        ws.cell(row=summary_row, column=4, value=round(total_ade / total_count, 3))
        ws.cell(row=summary_row, column=5, value=round(total_fde / total_count, 3))
        ws.cell(row=summary_row, column=6, value=f"共{total_count}条数据")
        
        # 调整列宽
        for col in range(1, len(headers)+1):
            ws.column_dimensions[chr(64+col)].width = 18
        
        # 保存文件
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"trajectory_evaluation_final_{timestamp}.xlsx"
        wb.save(filename)
        print(f"\n最终评估汇总已保存至: {filename}")

def main():
    rospy.init_node('simple_trajectory_evaluator')
    try:
        import openpyxl
    except ImportError:
        print("请先安装openpyxl库：pip install openpyxl")
        return
    
    evaluator = SimpleEvaluator()
    
    try:
        rospy.spin()
    except KeyboardInterrupt:
        print("\n程序正在结束，正在生成最终评估Excel...")
        evaluator.save_final_excel()  # 程序结束时保存Excel
        print("评估器已停止")

if __name__ == '__main__':
    main()