#!/usr/bin/env python3
import rospy
import math
import time
from datetime import datetime, timedelta
from visualization_msgs.msg import MarkerArray, Marker
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
# 缩短了更新的频率

class SimpleEvaluator:
    def __init__(self):
        # 存储数据 {障碍物ID: {意图类型: 轨迹点}}
        self.pred_trajectories = defaultdict(dict)
        self.history_trajectories = defaultdict(list)
        
        # 存储时间戳用于对齐
        self.last_pred_time = 0
        self.last_history_time = 0
        
        # 新增：累积所有评估数据（带时间戳）
        self.all_evaluation_data = []
        
        # 新增：时间窗参数（默认 None 表示不筛选，可通过ROS参数设置）
        self.eval_start_time = rospy.get_param("~eval_start", None)  # 格式："2024-05-20 15:30:00"
        self.eval_end_time = rospy.get_param("~eval_end", None)
        self.start_dt = datetime.strptime(self.eval_start_time, "%Y-%m-%d %H:%M:%S") if self.eval_start_time else None
        self.end_dt = datetime.strptime(self.eval_end_time, "%Y-%m-%d %H:%M:%S") if self.eval_end_time else None
        
        # 订阅话题
        self.pred_sub = rospy.Subscriber('/dynamic_predictor/predict_trajectories', 
                                        MarkerArray, self.pred_callback)
        self.history_sub = rospy.Subscriber('/dynamic_predictor/history_trajectories', 
                                           MarkerArray, self.history_callback)
        
        # 调整评估频率为0.1s（与意图预测频率同步）
        self.eval_interval = rospy.get_param("~eval_interval", 0.1)  # 可通过参数动态调整
        self.eval_timer = rospy.Timer(rospy.Duration(self.eval_interval), self.evaluate)
        
        print(f"简单轨迹评估器已启动，评估频率：{1/self.eval_interval}Hz，等待数据...")
        if self.start_dt and self.end_dt:
            print(f"仅统计时间窗：{self.eval_start_time} 至 {self.eval_end_time}")

    def pred_callback(self, msg):
        """预测轨迹回调（新增意图类型日志）"""
        self.last_pred_time = time.time()
        current_dt = datetime.now()
        
        for marker in msg.markers:
            # 提取障碍物ID和意图类型（关键：确认与发布方的ID编码规则一致）
            obstacle_id = marker.id // 10  # 发布方必须按此规则编码（如ID=12表示障碍物1，意图2）
            intent_type = marker.id % 10
            
            # 新增：打印意图信息，诊断是否有转弯意图（假设转弯意图对应特定数字，如1）
            if marker.type == Marker.LINE_STRIP:
                # 示例：如果意图类型为1代表转弯，此处打印提醒
                if intent_type == 1:  # 需根据实际定义修改
                    print(f"[{current_dt.strftime('%H:%M:%S')}] 检测到障碍物 {obstacle_id} 转弯意图（类型{intent_type}）")
                
                points = []
                for p in marker.points:
                    points.append([p.x, p.y, p.z])
                self.pred_trajectories[obstacle_id][intent_type] = points

    def history_callback(self, msg):
        """历史轨迹回调（不变）"""
        self.last_history_time = time.time()
        
        for marker in msg.markers:
            obstacle_id = marker.id
            
            if marker.type == Marker.LINE_STRIP:
                points = []
                for p in marker.points:
                    points.append([p.x, p.y, p.z])
                self.history_trajectories[obstacle_id] = points

    def calculate_distance(self, point1, point2):
        """计算两点间距离（不变）"""
        dx = point1[0] - point2[0]
        dy = point1[1] - point2[1]
        dz = point1[2] - point2[2]
        return math.sqrt(dx*dx + dy*dy + dz*dz)

    def calculate_ade_fde(self, pred_points, actual_points):
        """计算ADE和FDE（不变）"""
        if len(pred_points) == 0 or len(actual_points) == 0:
            return float('inf'), float('inf')
        
        min_length = min(len(pred_points), len(actual_points))
        total_error = 0
        for i in range(min_length):
            total_error += self.calculate_distance(pred_points[i], actual_points[i])
        ade = total_error / min_length
        fde = self.calculate_distance(pred_points[min_length-1], actual_points[min_length-1])
        
        return ade, fde

    def evaluate(self, event=None):
        """评估函数（新增时间窗筛选）"""
        current_time = time.time()
        current_dt = datetime.now()
        
        # 检查数据是否新鲜（0.2秒内，与0.1s频率匹配）
        if current_time - self.last_pred_time > 0.2 or current_time - self.last_history_time > 0.2:
            return
        
        # 新增：如果设置了时间窗，且当前不在窗口内，则不评估
        if self.start_dt and self.end_dt:
            if not (self.start_dt <= current_dt <= self.end_dt):
                return
        
        current_eval_data = []
        total_min_ade = 0
        total_min_fde = 0
        count = 0
        
        for obstacle_id in self.pred_trajectories:
            if obstacle_id in self.history_trajectories:
                actual_points = self.history_trajectories[obstacle_id]
                min_ade = float('inf')
                min_fde = float('inf')
                best_intent = -1
                
                # 新增：打印所有意图的ADE/FDE，判断是否有转弯意图的预测
                for intent_type, pred_points in self.pred_trajectories[obstacle_id].items():
                    ade, fde = self.calculate_ade_fde(pred_points, actual_points)
                    # 打印所有意图的误差（方便诊断）
                    print(f"障碍物 {obstacle_id} 意图{intent_type}：ADE={ade:.3f}, FDE={fde:.3f}")
                    
                    if ade < min_ade:
                        min_ade = ade
                        min_fde = fde
                        best_intent = intent_type
                
                if min_ade < float('inf'):
                    # 累积当前周期数据（带时间戳）
                    current_eval_data.append({
                        "timestamp": current_dt.strftime("%Y-%m-%d %H:%M:%S.%f"),  # 精确到微秒
                        "obstacle_id": obstacle_id,
                        "best_intent": best_intent,
                        "min_ade": min_ade,
                        "min_fde": min_fde,
                        "traj_length": len(actual_points)
                    })
                    
                    total_min_ade += min_ade
                    total_min_fde += min_fde
                    count += 1
        
        # 打印当前周期的平均值（可选，高频下可注释减少输出）
        if count > 0:
            avg_ade = total_min_ade / count
            avg_fde = total_min_fde / count
            print(f"\n[{current_dt.strftime('%H:%M:%S')}] 当前周期平均 (基于{count}个障碍物):")
            print(f"  平均minADE: {avg_ade:.3f}米，平均minFDE: {avg_fde:.3f}米")
        
        # 将当前周期数据添加到总列表
        self.all_evaluation_data.extend(current_eval_data)

    def save_final_excel(self):
        """保存Excel（仅保留时间窗内的数据）"""
        # 筛选时间窗内的数据
        filtered_data = []
        for data in self.all_evaluation_data:
            data_dt = datetime.strptime(data["timestamp"], "%Y-%m-%d %H:%M:%S.%f")
            if self.start_dt and self.end_dt:
                if self.start_dt <= data_dt <= self.end_dt:
                    filtered_data.append(data)
            else:
                filtered_data.append(data)
        
        if not filtered_data:
            print("\n时间窗内没有评估数据可保存")
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "轨迹评估汇总"
        
        # 设置表头
        headers = ["时间戳", "障碍物ID", "最佳意图类型", "minADE(米)", "minFDE(米)", "轨迹长度(点)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # 写入筛选后的数据
        for row, data in enumerate(filtered_data, 2):
            ws.cell(row=row, column=1, value=data["timestamp"])
            ws.cell(row=row, column=2, value=data["obstacle_id"])
            ws.cell(row=row, column=3, value=data["best_intent"])
            ws.cell(row=row, column=4, value=round(data["min_ade"], 3))
            ws.cell(row=row, column=5, value=round(data["min_fde"], 3))
            ws.cell(row=row, column=6, value=data["traj_length"])
        
        # 计算并写入总体平均值
        total_ade = sum(data["min_ade"] for data in filtered_data)
        total_fde = sum(data["min_fde"] for data in filtered_data)
        total_count = len(filtered_data)
        
        summary_row = total_count + 3
        cell = ws.cell(row=summary_row, column=1)
        cell.value = "时间窗内平均值"
        cell.font = Font(bold=True)
        
        ws.cell(row=summary_row, column=4, value=round(total_ade / total_count, 3))
        ws.cell(row=summary_row, column=5, value=round(total_fde / total_count, 3))
        ws.cell(row=summary_row, column=6, value=f"共{total_count}条数据")
        
        # 调整列宽
        for col in range(1, len(headers)+1):
            ws.column_dimensions[chr(64+col)].width = 20
        
        # 保存文件
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"trajectory_evaluation_window_{timestamp}.xlsx"
        wb.save(filename)
        print(f"\n时间窗评估结果已保存至: {filename}")

def main():
    rospy.init_node('simple_trajectory_evaluator')
    try:
        import openpyxl
    except ImportError:
        print("请先安装openpyxl库：pip install openpyxl")
        return
    
    evaluator = SimpleEvaluator()
    
    try:
        rospy.spin()
    except KeyboardInterrupt:
        print("\n程序正在结束，正在生成时间窗评估Excel...")
        evaluator.save_final_excel()
        print("评估器已停止")

if __name__ == '__main__':
    main()