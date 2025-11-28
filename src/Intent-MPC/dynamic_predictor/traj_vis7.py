#!/usr/bin/env python3
import rospy
import math
import time
import os
from datetime import datetime
from visualization_msgs.msg import MarkerArray, Marker
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font

class SimpleEvaluator:
    def __init__(self):
        self.pred_trajectories = defaultdict(dict)
        self.history_trajectories = defaultdict(list)
        self.last_pred_time = 0
        self.last_history_time = 0
        self.all_evaluation_data = []

        self.desktop_path = self.get_desktop_path()
        rospy.loginfo(f"检测到桌面路径: {self.desktop_path}")

        self.pred_sub = rospy.Subscriber('/dynamic_predictor/predict_trajectories',
                                         MarkerArray, self.pred_callback)
        self.history_sub = rospy.Subscriber('/dynamic_predictor/history_trajectories',
                                            MarkerArray, self.history_callback)

        self.eval_timer = rospy.Timer(rospy.Duration(1.0), self.evaluate)

        rospy.on_shutdown(self.save_final_excel)
        print("轨迹评估器已启动，按 Ctrl+C 时自动保存文件到桌面")

    def get_desktop_path(self):
        home = os.path.expanduser("~")
        chinese_desktop = os.path.join(home, "桌面")
        english_desktop = os.path.join(home, "Desktop")
        if os.path.exists(chinese_desktop):
            return chinese_desktop
        if os.path.exists(english_desktop):
            return english_desktop
        rospy.logwarn("未找到桌面目录，使用家目录")
        return home

    def pred_callback(self, msg):
        self.last_pred_time = time.time()
        for marker in msg.markers:
            obstacle_id = marker.id // 10
            intent_type = marker.id % 10
            if marker.type == Marker.LINE_STRIP:
                points = [[p.x, p.y, p.z] for p in marker.points]
                self.pred_trajectories[obstacle_id][intent_type] = points

    def history_callback(self, msg):
        self.last_history_time = time.time()
        for marker in msg.markers:
            obstacle_id = marker.id
            if marker.type == Marker.LINE_STRIP:
                points = [[p.x, p.y, p.z] for p in marker.points]
                self.history_trajectories[obstacle_id] = points

    def calculate_distance(self, p1, p2):
        return math.sqrt(
            (p1[0]-p2[0])**2 +
            (p1[1]-p2[1])**2 +
            (p1[2]-p2[2])**2
        )

    def calculate_ade_fde(self, pred_points, actual_points):
        if not pred_points or not actual_points:
            return float('inf'), float('inf')

        min_len = min(len(pred_points), len(actual_points))
        total_err = 0.0
        for i in range(min_len):
            total_err += self.calculate_distance(pred_points[i], actual_points[i])

        ade = total_err / min_len
        fde = self.calculate_distance(pred_points[min_len-1], actual_points[min_len-1])
        return ade, fde

    def evaluate(self, event=None):
        current_time = time.time()
        if current_time - self.last_pred_time > 2.0 or current_time - self.last_history_time > 2.0:
            return

        current_eval_data = []

        for obstacle_id in self.pred_trajectories:
            if obstacle_id not in self.history_trajectories:
                continue

            actual_points = self.history_trajectories[obstacle_id]

            # 找出最佳 ADE 的意图
            min_ade = float('inf')
            min_fde = float('inf')
            best_intent = -1
            intent_errors = {}

            # 计算所有意图的 ADE/FDE
            for intent_type, pred_points in self.pred_trajectories[obstacle_id].items():
                ade, fde = self.calculate_ade_fde(pred_points, actual_points)
                intent_errors[intent_type] = (ade, fde)

                if ade < min_ade:
                    min_ade = ade
                    min_fde = fde
                    best_intent = intent_type

            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")

            # 保存每一条意图的记录
            for intent_type, (ade, fde) in intent_errors.items():
                current_eval_data.append({
                    "timestamp": timestamp,
                    "obstacle_id": obstacle_id,
                    "intent_type": intent_type,
                    "ade": ade,
                    "fde": fde,
                    "is_best": 1 if intent_type == best_intent else 0,
                    "min_ade": min_ade,
                    "min_fde": min_fde,
                    "traj_length": len(actual_points)
                })

        self.all_evaluation_data.extend(current_eval_data)

    def save_final_excel(self):
        if not os.access(self.desktop_path, os.W_OK):
            rospy.logerr(f"目录不可写: {self.desktop_path}")
            self.desktop_path = "/tmp"
            rospy.logwarn(f"改为保存到: {self.desktop_path}")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"trajectory_evaluation_{timestamp}.xlsx"
        full_path = os.path.join(self.desktop_path, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "轨迹评估"

        # 完整表头
        headers = [
            "时间戳",
            "障碍物ID",
            "意图类型",
            "ADE(米)",
            "FDE(米)",
            "是否为最佳意图(1=是)",
            "minADE(米)",
            "minFDE(米)",
            "轨迹长度(点)"
        ]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = h
            cell.font = Font(bold=True)

        # 写入数据
        if self.all_evaluation_data:
            for row, data in enumerate(self.all_evaluation_data, 2):
                ws.cell(row=row, column=1, value=data["timestamp"])
                ws.cell(row=row, column=2, value=data["obstacle_id"])
                ws.cell(row=row, column=3, value=data["intent_type"])
                ws.cell(row=row, column=4, value=round(data["ade"], 3))
                ws.cell(row=row, column=5, value=round(data["fde"], 3))
                ws.cell(row=row, column=6, value=data["is_best"])
                ws.cell(row=row, column=7, value=round(data["min_ade"], 3))
                ws.cell(row=row, column=8, value=round(data["min_fde"], 3))
                ws.cell(row=row, column=9, value=data["traj_length"])
        else:
            ws.cell(row=2, column=1, value="无评估数据")
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

        try:
            wb.save(full_path)
            print(f"\n文件已保存到: {full_path}")
            print(f"你可以运行:  xdg-open {full_path}")
        except Exception as e:
            print(f"保存失败: {e}")
            fallback = os.path.abspath("trajectory_evaluation_backup.xlsx")
            wb.save(fallback)
            print(f"已保存到当前目录备用文件: {fallback}")


if __name__ == '__main__':
    rospy.init_node('trajectory_evaluator')
    evaluator = SimpleEvaluator()
    try:
        rospy.spin()
    except KeyboardInterrupt:
        evaluator.save_final_excel()
