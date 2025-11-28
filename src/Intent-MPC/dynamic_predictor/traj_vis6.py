#!/usr/bin/env python3
import rospy
import math
import time
import os
import sys
from datetime import datetime
from visualization_msgs.msg import MarkerArray, Marker
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# 可以新建文件到桌面，数据可能有点问题

class SimpleEvaluator:
    def __init__(self):
        self.pred_trajectories = defaultdict(dict)
        self.history_trajectories = defaultdict(list)
        self.last_pred_time = 0
        self.last_history_time = 0
        self.all_evaluation_data = []
        
        # 关键修复1：自动检测桌面路径（兼容中英文环境）
        self.desktop_path = self.get_desktop_path()
        rospy.loginfo(f"检测到桌面路径: {self.desktop_path}")
        
        # 订阅话题
        self.pred_sub = rospy.Subscriber('/dynamic_predictor/predict_trajectories', 
                                        MarkerArray, self.pred_callback)
        self.history_sub = rospy.Subscriber('/dynamic_predictor/history_trajectories', 
                                           MarkerArray, self.history_callback)
        
        # 定时评估
        self.eval_timer = rospy.Timer(rospy.Duration(1.0), self.evaluate)
        
        # 关键修复2：确保程序退出时保存文件（无论退出方式）
        rospy.on_shutdown(self.save_final_excel)
        print("轨迹评估器已启动，按Ctrl+C退出时自动保存结果到桌面")

    def get_desktop_path(self):
        """自动检测桌面路径（兼容中英文系统）"""
        home = os.path.expanduser("~")
        # 优先检测中文桌面路径
        chinese_desktop = os.path.join(home, "桌面")
        if os.path.exists(chinese_desktop) and os.path.isdir(chinese_desktop):
            return chinese_desktop
        # 再检测英文桌面路径
        english_desktop = os.path.join(home, "Desktop")
        if os.path.exists(english_desktop) and os.path.isdir(english_desktop):
            return english_desktop
        # 都不存在时使用家目录（保底方案）
        rospy.logwarn("未找到桌面目录，将使用家目录保存文件")
        return home

    def pred_callback(self, msg):
        self.last_pred_time = time.time()
        for marker in msg.markers:
            obstacle_id = marker.id // 10
            intent_type = marker.id % 10
            if marker.type == Marker.LINE_STRIP:
                points = [[p.x, p.y, p.z] for p in marker.points]
                self.pred_trajectories[obstacle_id][intent_type] = points

    def history_callback(self, msg):
        self.last_history_time = time.time()
        for marker in msg.markers:
            obstacle_id = marker.id
            if marker.type == Marker.LINE_STRIP:
                points = [[p.x, p.y, p.z] for p in marker.points]
                self.history_trajectories[obstacle_id] = points

    def calculate_distance(self, point1, point2):
        return math.sqrt((point1[0]-point2[0])**2 + (point1[1]-point2[1])** 2 + (point1[2]-point2[2])**2)

    def calculate_ade_fde(self, pred_points, actual_points):
        if not pred_points or not actual_points:
            return float('inf'), float('inf')
        min_length = min(len(pred_points), len(actual_points))
        total_error = sum(self.calculate_distance(pred_points[i], actual_points[i]) for i in range(min_length))
        return total_error/min_length, self.calculate_distance(pred_points[min_length-1], actual_points[min_length-1])

    def evaluate(self, event=None):
        current_time = time.time()
        if current_time - self.last_pred_time > 2.0 or current_time - self.last_history_time > 2.0:
            return
        
        current_eval_data = []
        for obstacle_id in self.pred_trajectories:
            if obstacle_id in self.history_trajectories:
                actual_points = self.history_trajectories[obstacle_id]
                min_ade = float('inf')
                min_fde = float('inf')
                best_intent = -1
                
                for intent_type, pred_points in self.pred_trajectories[obstacle_id].items():
                    ade, fde = self.calculate_ade_fde(pred_points, actual_points)
                    if ade < min_ade:
                        min_ade = ade
                        min_fde = fde
                        best_intent = intent_type
                
                if min_ade < float('inf'):
                    current_eval_data.append({
                        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f"),
                        "obstacle_id": obstacle_id,
                        "best_intent": best_intent,
                        "min_ade": min_ade,
                        "min_fde": min_fde,
                        "traj_length": len(actual_points)
                    })
        
        self.all_evaluation_data.extend(current_eval_data)

    def save_final_excel(self):
        """关键修复3：强制生成文件，即使数据为空也创建标记文件"""
        # 确保目录可写
        if not os.access(self.desktop_path, os.W_OK):
            rospy.logerr(f"错误：没有权限写入目录 {self.desktop_path}")
            # 退回到/tmp目录（Ubuntu下必定可写）
            self.desktop_path = "/tmp"
            rospy.logwarn(f"将使用临时目录保存：{self.desktop_path}")

        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"trajectory_evaluation_{timestamp}.xlsx"
        full_path = os.path.abspath(os.path.join(self.desktop_path, filename))

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "轨迹评估汇总"

        # 写入表头
        headers = ["时间戳", "障碍物ID", "最佳意图类型", "minADE(米)", "minFDE(米)", "轨迹长度(点)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)

        # 写入数据
        if self.all_evaluation_data:
            for row, data in enumerate(self.all_evaluation_data, 2):
                ws.cell(row=row, column=1, value=data["timestamp"])
                ws.cell(row=row, column=2, value=data["obstacle_id"])
                ws.cell(row=row, column=3, value=data["best_intent"])
                ws.cell(row=row, column=4, value=round(data["min_ade"], 3))
                ws.cell(row=row, column=5, value=round(data["min_fde"], 3))
                ws.cell(row=row, column=6, value=data["traj_length"])
        else:
            # 关键修复4：数据为空时明确标记，避免用户误解
            ws.cell(row=2, column=1, value="无评估数据")
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

        # 保存文件
        try:
            wb.save(full_path)
            print(f"\n文件已成功保存至：{full_path}")
            # 打印文件路径的终端命令，方便用户直接打开
            print(f"可在终端输入以下命令打开：xdg-open {full_path}")
        except Exception as e:
            print(f"\n保存文件失败：{str(e)}")
            # 最后的保底方案：保存到当前工作目录
            fallback_path = os.path.abspath(filename)
            wb.save(fallback_path)
            print(f"已尝试保存到当前目录：{fallback_path}")

if __name__ == '__main__':
    rospy.init_node('trajectory_evaluator')
    evaluator = SimpleEvaluator()
    try:
        rospy.spin()
    except KeyboardInterrupt:
        # 双重保障：即使rospy.on_shutdown失效，也手动调用保存
        evaluator.save_final_excel()